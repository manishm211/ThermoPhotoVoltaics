# This code performs the Monte Carlo uncertainty assessment on some input parameters and plots the probability distribution and cumulative distribution functions for LCOE (for an optimized system)

# %%
# This cell imports the required Python packages, Excel file (with data of power consumption and generation),
# defines the required arrays, and two dictionaries to store the input variables

import numpy as np
import warnings
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from distfit import distfit
import sys

my_generator = np.random.default_rng()
np.set_printoptions(threshold=sys.maxsize)

warnings.filterwarnings('ignore')            # To suppress warnings

import openpyxl 
wb = openpyxl.load_workbook('TPV.xlsx')
ws = wb['TPV']
Pg = [ws.cell(row=i,column=2).value for i in range(3,8763)]
Pc = [ws.cell(row=i,column=3).value for i in range(3,8763)]
Pg=np.array(Pg)
Pc=np.array(Pc)
Ch = np.zeros(8760)
Chh = [ws.cell(row=i,column=5).value for i in range(3,8763)]
Chc = [ws.cell(row=i,column=8).value for i in range(3,8763)]
Cc_max = ws.cell(row=8764,column=6).value
E_HTES = np.zeros(8761)
E_LTES = np.zeros(8761)
Pnet = np.zeros(8760)
P_loss_PV = np.zeros(8760)                   # Power lost from PV (energy is lost if HTES and LTES are at their max. capacity), W
P_in_LTES = np.zeros(8760)                   # Power input to LTES, W
P_in_HTES = np.zeros(8760)                   # Power input to HTES, W
P_grid = np.zeros(8760)                      # Power obtained from grid, W
P_out_PGU = np.zeros(8760)                   # Power output from PGU, W
Q_loss_PGU = np.zeros(8760)                  # Heat loss from PGU, Wh
Q_in_PGU = np.zeros(8760)                    # Heat input to PGU, Wh
Q_out_PGU = np.zeros(8760)                   # Heat output from PGU, Wh
Q_in_LTES = np.zeros(8760)                   # Low-grade exhaust heat sent to LTES from PGU, Wh
Q_loss_LTES = np.zeros(8761)                 # Heat loss for LTES, Wh
Q_loss_HTES = np.zeros(8761)                 # Heat loss for HTES, Wh
Qext = np.zeros(8760)                        # External heating from the external boiler, Wh
Q_out_LTES = np.zeros(8761)                  # Output heat from LTES, Wh
COP_EHP = 4                                  # COP of Electric Heat Pump (EHP) unit
COP_THP = 1.3                                # COP of THP unit
OPEX = np.zeros(8760)                        # Operating expenditures, defined in function "LCOE", $
Q_loss = np.zeros(8761)                      # Total heat loss (used to calculate self-consumption ratio), Wh

EUR_to_USD = 1.0775                          # To convert EUR to kWh, ref: https://www.google.com/finance/quote/EUR-USD?sa=X&ved=2ahUKEwjr-5i7opyEAxXhj4kEHTtdCtQQmY0JegQIBxAv \
# for Date- 8 Feb, 17:28:00 UTC

a_LTES = 0.1                                 # LTES heat loss parameters, WK^−1 dm^−3/2
V_LTES = 20/0.08                             # Water volume in L (or dm^3)=20 kWh_th/energy storage density (=0.08 kWh_th/L)
a_HTES = 0.1                                 # HTES heat loss parameters, WK^−1 dm^−3/2
T_HTES = 1414                                # Operational temperature of storage media, K
delta_T_LTES_E = 70                          # Mean temp b/w LTES and ambient, degree C, for PHPS-E unit
#delta_T_LTES_T = 130                        # Mean temp b/w LTES and ambient, degree C, for PHPS-T unit
delta_T_HTES = 1200                          # Mean temp b/w HTES and ambient, degree C
Ed_HTES = 6.74*pow(10,-7)*pow(T_HTES,2)-1.72*pow(10,-4)*T_HTES+0.140 #HTES energy storage density, kWh_th/L

Model = {
    "E_HTES_max": 15,                        # HTES maximum capacity, kWh_th
    "E_LTES_max": 80,                        # LTES maximum capacity, kWh_th
    "deltat": 1,                             # Time interval, hours
    "Eff_PGU": 0.2745,                       # Average TPV efficiency
    "Pmax_PGU": 0.5,                         # PGU maximum generation capacity, kWel
    "Lifetime": 24.34,                       # Lifetime of the project, years
    "CAPEX*_HTES": 118.4,                    # HTES Capital cost factor, $/kWh_th
    "CAPEX*_LTES": 30*EUR_to_USD,            # LTES Capital cost factor, $/kWh_th
    "CAPEX*_PGU": 1184.66,                   # PGU Capital cost factor, $/kW_el
    "P_nom_PV": 5,                           # Nominal PV power, kWp
    "CAPEX*_PV": 5163.74,                    # PV Capital cost factor, $/kW_el
    "CAPEX*_EHP": 500*EUR_to_USD,            # EHP Capital cost factor, $/kW_cool
    #"CAPEX*_THP": 500*EUR_to_USD,           # THP Capital cost factor, $/kW_cool
    "OPEX_elec_var": 0.124,                  # Variable cost of grid electricity, $/kWh_el
    "OPEX_elec_fix": 50*EUR_to_USD,          # Fixed cost of grid electricity, $/kWh_el-year
    "OPEX_fuel_var": 0.015,                  # Variable cost of fuel, $/kWh_th
    "OPEX_fuel_fix": 60*EUR_to_USD,          # Fixed cost of fuel, $/kWh_el-year
    "WACC_nom": 0.0217,                      # Nominal WACC (Weighted average cost of capital)
    "WACC_real": 0,                          # Real WACC
    "Infl_overall": 0.02,                    # Overall inflation
    "Infl_e": 0.0231,                        # Inflation for fuel and electricity
    "LCOE": 0,                               # Levelized Cost of Energy, $/kWh
    "LCOE_el": 0,                            # Levelized Cost of Electricity, $/kWh
    "SCR": 0,                                # Self-consumption ratio
    "OPEX_el": 0,                            # Operating expenditures for LCOE_el
    "OPEX": 0,                               # Operating expenditures for LCOE
    "Den_LCOE": 0,                           # Denominator of LCOE
    "Den_LCOE_el": 0,                        # Denominator of LCOE_el
    "Q_ext_total": 0,                        # Total energy obtained from external boiler
    "P_grid_total": 0,                       # Total power obtained from grid
    "P_PV_total": 0,                         # Total power obtained from solar PV
    "Emission_factor_natural_gas": 180.7975, # Emission factor for natural gas, gCO2/kWh
    "Emission_factor_grid_elec": 304.14,     # Emission factor for grid electricity, gCO2eq/kWh
    "Emission_factor_PV_elec": 43,           # Emission factor for PV electricity, gCO2eq/kWh
    "Emissions_total_natural_gas": 0,        # Total emissions from natural gas, gCO2
    "Emissions_total_grid_elec": 0,          # Total emissions from grid electricity, gCO2eq
    "Emissions_total_PV_elec": 0,            # Total emissions from PV electricity, gCO2eq
    }

# The dictionary below is defined for changing the values to default
sens_Model={
    "CAPEX*_HTES": 118.4,         
    "CAPEX*_PGU": 1184.66,        
    "CAPEX*_PV": 5163.74,         
    "Infl_e": 0.0231,             
    "WACC_nom": 0.0217,           
}

Q_loss_HTES= np.full(8760, ((a_HTES*np.sqrt(Model["E_HTES_max"]/Ed_HTES)*delta_T_HTES)/1000))  # HTES energy losses for a time interval
Q_loss_LTES = np.full(8760, ((a_LTES*np.sqrt(Model["E_LTES_max"]/0.018)*delta_T_LTES_E)/1000)) # LTES energy losses for a time interval
Pnet = Pc - Pg                                                                                 # The difference between power consumption and generation for a time interval

#%%
# This cell contains the function that follows the energy management algorithm to calculate LCOE

# Function to calcualte LCOE
def Monte_Carlo(OPEX_fuel_var,Infl_e,Eff_PGU,Lifetime,CAPEX_HTES,CAPEX_PGU,CAPEX_PV,OPEX_elec_var,WACC_nom):

        Model["OPEX_fuel_var"] = OPEX_fuel_var
        Model["Infl_e"] = Infl_e
        Model["Eff_PGU"] = Eff_PGU
        Model["Lifetime"] = Lifetime
        Model["CAPEX*_HTES"] = CAPEX_HTES
        Model["CAPEX*_PGU"] = CAPEX_PGU
        Model["CAPEX*_PV"] = CAPEX_PV
        Model["OPEX_elec_var"] = OPEX_elec_var
        Model["WACC_nom"] = WACC_nom
        # The above lines store the values of the Monte Carlo parameters that are supplied one by one using their fitted probability distribution

        Model["WACC_real"] = (1 + Model["WACC_nom"])/(1 + Model["Infl_overall"]) - 1
        # Formula to compute Real WACC
        x = 0.99
        y = 0.1
        # x and y represent the charging condition for HTES and LTES required to evaluate certain parameters in the code
        for i in range(8760):
            
            if Pnet[i]<0:
                # if Pnet is negative (power gen. > cons.), no power is obtained from grid (assuming PV energy can be directly supplied to the load or stored in HTES or LTES)
                P_grid[i] = 0
                P_out_PGU[i] = 0
                Q_loss_PGU[i] = 0
                Q_in_PGU[i] = 0
                Q_out_PGU[i] = 0
                Q_in_LTES[i] = 0
                if E_HTES[i]<Model["E_HTES_max"]:
                    # HTES charging is prioritized over LTES: here, it is checked if it is fully charged or not
                    if Chh[i]>=0:
                        # to check if Heat consumption>0                                                    
                        if E_HTES[i]>x*Model["E_HTES_max"]:
                            # to check if HTES charge is higher than 99%                         
                            P_in_HTES[i]=0
                            # Power input to HTES = 0 as HTES is near to its capacity                                          
                            if E_LTES[i]<y*Model["E_LTES_max"]:
                                # to check if LTES charge is lower than 10%                     
                                P_loss_PV[i]=0
                                P_in_LTES[i]=-Pnet[i]
                                # LTES is charged if HTES is almost fully charged, and LTES if charges less than 10%
                            else:
                                P_loss_PV[i]=0
                                # Lost PV electricity = 0                                      
                                P_in_LTES[i]=0
                                # Input power to LTES = 0 as it is almost fully charged
                        else:
                            P_in_HTES[i]=-Pnet[i]
                            # Power input to HTES is positive (as it is not fully charged) and equal to Pnet
                            P_loss_PV[i]=0
                            # Power lost from PV = 0
                            P_in_LTES[i]=0
                            # No power goes to LTES
                    else:
                        P_in_HTES[i]=-Pnet[i]
                        P_loss_PV[i]=0
                        P_in_LTES[i]=0
                else:                                                               
                    P_in_HTES[i]=0
                    # HTES is at its max. capacity, thus it cannot store anymore
                    if E_LTES[i]>Model["E_LTES_max"]:
                        # to check is LTES is at its max capacity, thus it cannot store anymore, thus the extra energy produced by PV is lost                               
                        P_loss_PV[i]=-Pnet[i] 
                        # Lost PV electricity = Pnet, as both HTES and LTES are fully charged                                      
                        P_in_LTES[i]=0
                        # Power input to LTES = 0
                    else:
                        P_loss_PV[i]=0 
                        # Power lost from PV = 0                                            
                        P_in_LTES[i]=-Pnet[i]
                        # Power input to LTES = Pnet, as it is not fully charged
            else:     
                # Pnet is greater than or equal to zero (energy cons. >= gen.)                                                              
                P_loss_PV[i] = 0
                # Lost PV electricity = 0, thus, all of it goes to supply the electrical load
                P_in_LTES[i] = 0
                # LTES input power = 0
                P_in_HTES[i] = 0
                # HTES input power = 0
                
                if E_HTES[i]>Pnet[i]*Model["deltat"]/Model["Eff_PGU"]:
                    # To check if HTES has enough stored energy to supply             
                    if Pnet[i]>Model["Pmax_PGU"]:
                        # To check if the electricity demand is higher than what PGU can supply                                  
                        P_grid[i]=Pnet[i]-Model["Pmax_PGU"]
                        # If higher, take what you can from the grid
                        P_out_PGU[i]=Model["Pmax_PGU"]
                        # The rest of electricity is supplied from PGU
                    else:
                        P_grid[i]=0
                        # No power is obtained from the grid as PGU can supply all that is required
                        P_out_PGU[i]=Pnet[i]
                        # PGU supplies Pnet
                else:
                    # HTES does not have enough stored energy
                    P_grid[i]=Pnet[i]
                    # Grid supplies Pnet
                    P_out_PGU[i]=0
                    # No power is obtained from PGU
                Q_in_PGU[i]=P_out_PGU[i]/Model["Eff_PGU"]
                # Heat input to PGU ius calculated based on output and efficiency
                Q_out_PGU[i]=Q_in_PGU[i]-P_out_PGU[i]
                # Energy balance on PGU
                if E_LTES[i]>Model["E_LTES_max"]:
                    # To check if LTES is at its max. capacity, thus it cannot store anymore                                  
                    Q_loss_PGU[i]=Q_out_PGU[i]
                    # Heat output of PGU is lost and not sent to LTES
                    Q_in_LTES[i]=0
                    # Input power to LTES = 0
                else:
                    Q_loss_PGU[i]=0
                    # Power lost from PGU = 0
                    Q_in_LTES[i]=Q_out_PGU[i]
                    # PGU output goes to LTES as it is not fully charged
            
            Ch[i] = Chh[i]                                                         
            #Ch[i] = Chh[i] + Chc[i]                                               
            if Ch[i]<((E_LTES[i]/Model["deltat"])+Q_in_LTES[i]+P_in_LTES[i]-Q_loss_LTES[i]):
                # To check if the energy stored in LTES is enough to supply the heating load
                Qext[i]=0
                # Heat obtained from the external boiler = 0, as LTES can satisfy the heating requirement
                Q_out_LTES[i]=Ch[i]
                # LTES output energy = the heating requirement        
            else:
                # LTES cannot supply the heating load
                Q_out_LTES[i] = E_LTES[i]/Model["deltat"]+Q_in_LTES[i]+P_in_LTES[i]-Q_loss_LTES[i]
                # Energy balance at LTES to calculate energy output
                if Q_out_LTES[i]<0:
                    Q_out_LTES[i]=0
                    # To prevent LTES output from becoming negative
                Qext[i] = Ch[i]-Q_out_LTES[i]
                # Rest of the heating requirement is obtained from the external boiler

            if (E_LTES[i]+(P_in_LTES[i]+Q_in_LTES[i]-Q_out_LTES[i]-Q_loss_LTES[i])*Model["deltat"])<0:
                # If the above inequality is satisfied, it means that the LTES is unable to store energy for 
                # the ith interval, thus take the ith interval's energy from the external boiler, and keep
                # the ith interval unused LTES energy stored for the next, (i+1)th interval.
                E_LTES[i+1]=E_LTES[i]
                P_in_LTES[i] = 0
                Q_in_LTES[i] = 0
                Q_out_LTES[i] = 0
                Q_loss_LTES[i] = 0
                Qext[i] = Ch[i]

            else:
                E_LTES[i+1] = E_LTES[i]+(P_in_LTES[i]+Q_in_LTES[i]-Q_out_LTES[i]-Q_loss_LTES[i])*Model["deltat"]

            if (E_HTES[i]+(P_in_HTES[i]-Q_in_PGU[i]-Q_loss_HTES[i])*Model["deltat"])<0:
                #If the above inequality is satisfied, it means that the HTES is unable to store energy for 
                #the ith interval, thus take the ith interval's energy from the grid, and keep
                #the ith interval unused HTES energy stored for the next, (i+1)th interval.
                E_HTES[i+1]=E_HTES[i]
                #Q_loss_HTES[i]=0
                #Q_loss_HTES[i]=0
                P_in_HTES[i] = 0
                Q_in_PGU[i] = 0
                Q_loss_HTES[i] = 0
                Q_out_PGU[i] = 0
                P_out_PGU[i] = 0
                Q_loss_PGU[i] = 0

                P_grid[i] = Pc[i] - ( (Pg[i] - (P_in_LTES[i]+P_loss_PV[i]+P_in_HTES[i])) + P_out_PGU[i])
                # Energy balance at LTES to calculate the stored energy for the (i+1)th interval
                
                if P_grid[i]<0:
                    # To check if Pgrid is negative
                    P_loss_PV[i] = ( (Pg[i] - (P_in_LTES[i]+Pc[i]+P_in_HTES[i])) + P_out_PGU[i])
                    # Rest of the electricity is obtained from solar PV
                    P_grid[i] = 0
                    # Pgrid is turned zero if it were negative

                Q_in_LTES[i] = 0
                # Heat input to LTEs = 0

            else:
                E_HTES[i+1] = E_HTES[i]+(P_in_HTES[i]-Q_in_PGU[i]-Q_loss_HTES[i])*Model["deltat"]    
                # Energy balance at HTES to calculate the stored energy for the (i+1)th interval

            Q_loss = Q_loss_HTES+Q_loss_LTES+Q_loss_PGU+P_loss_PV
            # Total energy losses are computed as a sum of losses from HTEs, LTES, PGU, and PV

        Model["CAPEX"] = Model["P_nom_PV"]*Model["CAPEX*_PV"] + Model["E_HTES_max"]*Model["CAPEX*_HTES"] + Model["E_LTES_max"]*Model["CAPEX*_LTES"]\
        + Model["Pmax_PGU"]*Model["CAPEX*_PGU"] + (Cc_max/COP_EHP)*Model["CAPEX*_EHP"]\
        # Total capital expenditure, in $ is computed as the sum of the product of capital cost factors and energy capacities of PV, HTES, LTE, PGU, and EHP

        Term_1 = np.sum(P_grid)
        # Total electricity obtained from grid throughout the year
        Term_2 = np.sum(Pc)
        # Total electricity consumed by load throughout the year
        Term_3 = np.sum(Qext)
        # Total heat obtained from external boiler (running on natural gas) throughout the year
        Term_4 = np.sum(Pc+Ch)
        # Total energy requirements (heat + electricity) throughout the year

        def func_OPEX(x):
            # This function calculates the operating expenditure (normalized by WACC_nom) for LCOE_el
            # It consists of fixed and variable cost of grid electricity (both multiplied by electricity inflation rate)
            return ((pow((1+Model["Infl_e"]),x)*(Model["OPEX_elec_fix"]*max(P_grid)+Model["OPEX_elec_var"]*Term_1)/(pow((1+Model["WACC_nom"]),x))))
        
        def func_OPEX_2(x):
            # This function calculates the operating expenditure (normalized by WACC_nom) for LCOE
            # It consists of fixed and variable cost of grid electricity and natural gas (both multiplied by electricity and fuel inflation rate)
            return ((pow((1+Model["Infl_e"]),x)*(Model["OPEX_elec_fix"]*max(P_grid)+Model["OPEX_elec_var"]*Term_1)/(pow((1+Model["WACC_nom"]),x))+\
            pow((1+Model["Infl_e"]),x)*(Model["OPEX_fuel_fix"]+Model["OPEX_fuel_var"]*Term_3)/(pow((1+Model["WACC_nom"]),x))))
        
        def func_Den(x):
            # This function calculates the denominator of LCOE: yearly electric demand normalized by WACC_real
            return (Term_2/(pow((1+Model["WACC_real"]),x)))
        
        def func_Den_2(x):
            # This function calculates the denominator of LCOE: yearly energy (heat + electricity) demand normalized by WACC_real
            return (Term_4/(pow((1+Model["WACC_real"]),x)))
        
        L = np.arange(1,Model["Lifetime"]+1,1)

        Model["OPEX_el"]=np.sum(func_OPEX(L))
        # This function calculates the operating expenditures (for LCOE_el) for the entire lifetime of the project
        Model["LCOE_el"] = (Model["CAPEX"] + Model["OPEX_el"])/(np.sum(func_Den(L)))
        # LCOE_el ($/kWh) is computed according to its definition
        Model["OPEX"] = np.sum(func_OPEX_2(L))
        # This function calculates the operating expenditures (for LCOE) for the entire lifetime of the project
        Model["LCOE"] = (Model["CAPEX"] + Model["OPEX"])/(np.sum(func_Den_2(L)))
        # LCOE ($/kWh) is computed according to its definition

        Model["WACC_nom"] = sens_Model["WACC_nom"]
        Model["Infl_e"] = sens_Model["Infl_e"]
        Model["CAPEX*_HTES"] = sens_Model["CAPEX*_HTES"]
        Model["CAPEX*_PGU"] = sens_Model["CAPEX*_PGU"]
        Model["CAPEX*_PV"] = sens_Model["CAPEX*_PV"]
        #Setting the variables to their default values
        return Model["LCOE"]

# %%
# This cell imports an Excel file with historical data on the non-Triangular distribution parameters, fits a probability distribution for all parameters, and stores random numbers extracted from them in a dataframe
# The fitted probability distributions for all parameters are already identified from their historical values

sns.set_style('whitegrid')
num_reps = 1000                      # Number of repetitions/random numbers generated from the Fitted probability distributions
num_simulations = 100                 # Number of simulations
wb = openpyxl.load_workbook('Monte_Carlo_parameters.xlsx')

ws = wb['Natural_gas']
Natural_gas = [ws.cell(row=i,column=2).value for i in range(4,1002)]
Natural_gas=np.array(Natural_gas)
Natural_gas_dist = distfit(distr='genextreme')
Natural_gas_dist.fit_transform(Natural_gas)
Natural_gas_dist.plot(chart='pdf', fontsize=40, xlabel="Natural gas price, $/kWh (Genextreme: 0.011, 0.004)", ylabel="Frequency")
plt.legend(["Values","Best-fit","95% confidence interval"], fontsize = 36, loc=0)
Natural_gas_dist_gen = Natural_gas_dist.generate(num_reps)

ws = wb['Inflation_rate']
Inflation_rate = [ws.cell(row=i,column=2).value for i in range(2,290)]
Inflation_rate=np.array(Inflation_rate)
Inflation_rate_dist = distfit(distr='t')
Inflation_rate_dist.fit_transform(Inflation_rate)
Inflation_rate_dist.plot(chart='pdf', fontsize=40, xlabel="Inflation rate, % (t: 2.314, 1.137)", ylabel="Frequency")
plt.legend(["Values","Best-fit","95% confidence interval"], fontsize = 36, loc=0)
Inflation_rate_dist_gen = Inflation_rate_dist.generate(num_reps)

PGU_Efficiency = my_generator.triangular(16.5,25,41,num_reps)
PGU_Efficiency_dist = distfit(distr='triang')
PGU_Efficiency_dist.fit_transform(PGU_Efficiency)
PGU_Efficiency_dist.plot(chart='pdf', fontsize=40, xlabel="PGU efficiency, % (Beta: 16.500, 39.856)", ylabel="Frequency")
plt.legend(["Values","Best-fit","95% confidence interval"], fontsize = 36, loc=0)
PGU_Efficiency_dist_gen = PGU_Efficiency_dist.generate(num_reps)

Lifetime = my_generator.triangular(18,25,30,num_reps)
Lifetime_dist = distfit(distr='triang')
Lifetime_dist.fit_transform(Lifetime)
Lifetime_dist.plot(chart='pdf', fontsize=40, title="Best-fit distribution for Lifetime", xlabel="Lifetime (years)", ylabel="Frequency")
Lifetime_dist_gen = PGU_Efficiency_dist.generate(num_reps)

CAPEX_HTES = my_generator.triangular(30*EUR_to_USD,100*EUR_to_USD,200*EUR_to_USD,num_reps)
CAPEX_HTES_dist = distfit(distr='triang')
CAPEX_HTES_dist.fit_transform(CAPEX_HTES)
CAPEX_HTES_dist.plot(chart='pdf', fontsize=40, title="Best-fit distribution for HTES CAPEX", xlabel="HTES CAPEX ($/kWhth)", ylabel="Frequency")
CAPEX_HTES_dist_gen = CAPEX_HTES_dist.generate(num_reps)

CAPEX_PGU = my_generator.triangular(300*EUR_to_USD,1000*EUR_to_USD,2000*EUR_to_USD,num_reps)
CAPEX_PGU_dist = distfit(distr='triang')
CAPEX_PGU_dist.fit_transform(CAPEX_PGU)
CAPEX_PGU_dist.plot(chart='pdf', fontsize=40, title="Best-fit distribution for PGU CAPEX", xlabel="PGU CAPEX ($/kW)", ylabel="Frequency")
CAPEX_PGU_dist_gen = CAPEX_PGU_dist.generate(num_reps)

ws = wb['CAPEX_PV']
CAPEX_PV = [ws.cell(row=6,column=i).value for i in range(4,14)]
CAPEX_PV=np.array(CAPEX_PV)
CAPEX_PV_dist = distfit(distr='expon')
CAPEX_PV_dist.fit_transform(CAPEX_PV)
CAPEX_PV_dist.plot(chart='pdf', fontsize=40, xlabel="PV CAPEX, $/kW (Exponential: 3820.849,1342.887)", ylabel="Frequency")
plt.legend(["Values","Best-fit","95% confidence interval"], fontsize = 36, loc=0)
CAPEX_PV_dist_gen = CAPEX_PV_dist.generate(num_reps)

ws = wb['Electricity_price']
Electricity_price = [ws.cell(row=i,column=2).value for i in range(6,233)]
Electricity_price=np.array(Electricity_price)
Electricity_price_dist = distfit(distr='dweibull')
Electricity_price_dist.fit_transform(Electricity_price)
Electricity_price_dist.plot(chart='pdf', fontsize=40, xlabel="Electricity price, $/kWh (Weibull: 0.124, 0.012)", ylabel="Frequency")
plt.legend(["Values","Best-fit","95% confidence interval"], fontsize = 36, loc=0)
Electricity_price_dist_gen = Electricity_price_dist.generate(num_reps)

WACC_nominal = my_generator.triangular(1.5,2,3, num_reps)
WACC_nominal_dist = distfit(distr='triang')
WACC_nominal_dist.fit_transform(WACC_nominal)
WACC_nominal_dist.plot(chart='pdf', fontsize=40, title="Best-fit distribution for nominal WACC", xlabel="Nominal WACC (%)", ylabel="Frequency")
WACC_nominal_dist_gen = WACC_nominal_dist.generate(num_reps)

df = pd.DataFrame(index=range(num_reps), data={'Natural_gas': Natural_gas_dist_gen,
                                               'Inflation_rate': Inflation_rate_dist_gen/100,
                                               'PGU_Efficiency': PGU_Efficiency/100,
                                               'Lifetime': Lifetime,
                                               'CAPEX_HTES': CAPEX_HTES,
                                               'CAPEX_PGU': CAPEX_PGU,
                                               'CAPEX_PV': CAPEX_PV_dist_gen,
                                               'Electricity_price': Electricity_price_dist_gen,
                                               'WACC_nominal': WACC_nominal/100,
                                               })

# %%
# This cell calculates the LCOE for the Monte Carlo parameters' values, and plots a fitted probability and cumulative distribution function with 90% confidence interval

for i in range(num_simulations):
    df['LCOE'] = df.apply(lambda x: Monte_Carlo(x['Natural_gas'],x['Inflation_rate'],x['PGU_Efficiency'],x['Lifetime'],x['CAPEX_HTES'],x['CAPEX_PGU'],x['CAPEX_PV'],x['Electricity_price'],x['WACC_nominal']), axis=1)

LCOE=np.array(df['LCOE'])
LCOE_dist = distfit(alpha=0.1)
LCOE_dist.fit_transform(LCOE)
LCOE_dist.plot(chart='pdf',emp_properties=None, pdf_properties={'color': 'k', 'linewidth': 7, 'linestyle': 'dashed' }, fontsize=40, xlabel="LCOE ($/kWh)", ylabel="Frequency")
plt.title("Best-fit distribution for LCOE",fontsize=42)
plt.legend(["Best-fit","90% confidence interval"], fontsize = 36, loc=0)
plt.xlim(left=0.00,right=0.20)
plt.savefig('Monte_Carlo_LCOE_PDF_optimized_system.jpeg', dpi=300, bbox_inches='tight')
dfit = distfit(alpha=0.1)
dfit.fit_transform(df['LCOE'])

fig, ax = dfit.plot(chart='pdf', fontsize=40, title="Probability Distribution Function (PDF)", xlabel="LCOE ($/kWh)", ylabel="Frequency")
ax.set_xlim(left=0.00,right=0.20)
fig, ax = dfit.plot(chart='cdf', fontsize=40, xlabel="LCOE ($/kWh)", ylabel="Frequency")
ax.set_xlim(left=0.00,right=0.20)
plt.title("Cumulative Distribution Function (CDF)",fontsize=42)
plt.legend(["Empirical CDF","Best-fit","90% confidence interval"], fontsize=36, loc=0)
plt.savefig('Monte_Carlo_LCOE_CDF_optimized_system.jpeg', dpi=300, bbox_inches='tight')
dfit.plot(chart='pdf', pdf_properties={'color': 'r'}, cii_properties={'color': 'g'}, emp_properties=None, bar_properties=None, title="Probability Distribution Function (PDF)", fontsize=40,xlabel="LCOE ($/kWh)", ylabel="Frequency")
dfit.plot(chart='cdf', pdf_properties={'color': 'r'}, cii_properties={'color': 'g'}, emp_properties=None, bar_properties=None, title="Cumulative Distribution Function (CDF)", fontsize=40,xlabel="LCOE ($/kWh)", ylabel="Frequency")

fig, ax = plt.subplots(1,2, figsize=(25, 10))
dfit.plot(chart='pdf', ax=ax[0], title="Probability Distribution Function (PDF)", xlabel="LCOE ($/kWh)", ylabel="Frequency")
dfit.plot(chart='cdf', ax=ax[1], title="Cumulative Distribution Function (CDF)", xlabel="LCOE ($/kWh)", ylabel="Frequency")
fig, ax = dfit.plot(chart='pdf', pdf_properties={'color': 'r', 'linewidth': 3}, cii_properties={'color': 'r', 'linewidth': 3}, bar_properties={'color': '#1e3f5a'}, title="Probability Distribution Function (PDF)", fontsize=40,xlabel="LCOE ($/kWh)", ylabel="Frequency")
dfit.plot(chart='cdf', n_top=10, pdf_properties={'color': 'r'}, cii_properties=None, bar_properties=None, ax=ax, title="Probability Distribution Function (PDF)", fontsize=40,xlabel="LCOE ($/kWh)", ylabel="Frequency")