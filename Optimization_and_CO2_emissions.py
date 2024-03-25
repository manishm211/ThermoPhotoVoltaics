# This code calculates and optimizes LCOE and LCOE_el, plots them, calculates the preliminary CO2eq. emissions
# for these scenarios- no energy storage, base-case system, and optimized system, and plots them.

# %%
# This cell imports the required Python packages, Excel file (with data of power consumption and generation),
# defines the required arrays, and a dictionary to store the input variables

import numpy as np
import warnings
import matplotlib.pyplot as plt
from scipy.optimize import minimize

warnings.filterwarnings('ignore')            # To suppress warnings

import openpyxl 
wb = openpyxl.load_workbook('TPV.xlsx')
ws = wb['TPV']
Pg = [ws.cell(row=i,column=2).value for i in range(3,8763)]
Pc = [ws.cell(row=i,column=3).value for i in range(3,8763)]

Ch = np.zeros(8760)
Chh = [ws.cell(row=i,column=5).value for i in range(3,8763)]
Chc = [ws.cell(row=i,column=8).value for i in range(3,8763)]
Cc_max = ws.cell(row=8764,column=6).value
E_HTES = np.zeros(8761)
E_LTES = np.zeros(8761)
Pnet = np.zeros(8760)
P_loss_PV = np.zeros(8760)                   # Power lost from PV (energy is lost if HTES and LTES are at their max. capacity), W
P_in_LTES = np.zeros(8760)                   # Power input to LTES, W
P_in_HTES = np.zeros(8760)                   # Power input to HTES, W
P_grid = np.zeros(8760)                      # Power obtained from grid, W
P_out_PGU = np.zeros(8760)                   # Power output from PGU, W
Q_loss_PGU = np.zeros(8760)                  # Heat loss from PGU, Wh
Q_in_PGU = np.zeros(8760)                    # Heat input to PGU, Wh
Q_out_PGU = np.zeros(8760)                   # Heat output from PGU, Wh
Q_in_LTES = np.zeros(8760)                   # Low-grade exhaust heat sent to LTES from PGU, Wh
Q_loss_LTES = np.zeros(8761)                 # Heat loss for LTES, Wh
Q_loss_HTES = np.zeros(8761)                 # Heat loss for HTES, Wh
Qext = np.zeros(8760)                        # External heating from the external boiler, Wh
Q_out_LTES = np.zeros(8761)                  # Output heat from LTES, Wh
COP_EHP = 4                                  # COP of Electric Heat Pump (EHP) unit
COP_THP = 1.3                                # COP of THP unit
OPEX = np.zeros(8760)                        # Operating expenditures, defined in function "LCOE", $
Q_loss = np.zeros(8761)                      # Total heat loss (used to calculate self-consumption ratio), Wh

EUR_to_USD = 1.0775                          # To convert EUR to kWh, ref: https://www.google.com/finance/quote/EUR-USD?sa=X&ved=2ahUKEwjr-5i7opyEAxXhj4kEHTtdCtQQmY0JegQIBxAv \
# for Date- 8 Feb, 17:28:00 UTC

a_LTES = 0.1                                 # LTES heat loss parameters, WK^−1 dm^−3/2
V_LTES = 20/0.08                             # Water volume in L (or dm^3)=20 kWh_th/energy storage density (=0.08 kWh_th/L)
a_HTES = 0.1                                 # HTES heat loss parameters, WK^−1 dm^−3/2
T_HTES = 1414                                # Operational temperature of storage media, K
delta_T_LTES_E = 70                          # Mean temp b/w LTES and ambient, degree C, for PHPS-E unit
#delta_T_LTES_T = 130                        # Mean temp b/w LTES and ambient, degree C, for PHPS-T unit
delta_T_HTES = 1200                          # Mean temp b/w HTES and ambient, degree C
Ed_HTES = 6.74*pow(10,-7)*pow(T_HTES,2)-1.72*pow(10,-4)*T_HTES+0.140 #HTES energy storage density, kWh_th/L

Model = {
    "E_HTES_max": 40,                        # HTES maximum capacity, kWh_th
    "E_LTES_max": 80,                        # LTES maximum capacity, kWh_th
    "deltat": 1,                             # Time interval, hours
    "Eff_PGU": 0.2745,                       # Average TPV efficiency
    "Pmax_PGU": 1,                           # PGU maximum generation capacity, kWel
    "Lifetime": 24.34,                       # Lifetime of the project, years
    "CAPEX*_HTES": 118.4,                    # HTES Capital cost factor, $/kWh_th
    "CAPEX*_LTES": 30*EUR_to_USD,            # LTES Capital cost factor, $/kWh_th
    "CAPEX*_PGU": 1184.66,                   # PGU Capital cost factor, $/kW_el
    "P_nom_PV": 9.5,                         # Nominal PV power, kWp
    "CAPEX*_PV": 5163.74,                    # PV Capital cost factor, $/kW_el
    "CAPEX*_EHP": 500*EUR_to_USD,            # EHP Capital cost factor, $/kW_cool
    #"CAPEX*_THP": 500*EUR_to_USD,           # THP Capital cost factor, $/kW_cool
    "OPEX_elec_var": 0.124,                  # Variable cost of grid electricity, $/kWh_el
    "OPEX_elec_fix": 50*EUR_to_USD,          # Fixed cost of grid electricity, $/kWh_el-year
    "OPEX_fuel_var": 0.015,                  # Variable cost of fuel, $/kWh_th
    "OPEX_fuel_fix": 60*EUR_to_USD,          # Fixed cost of fuel, $/kWh_el-year
    "WACC_nom": 0.0217,                      # Nominal WACC (Weighted average cost of capital)
    "WACC_real": 0,                          # Real WACC
    "Infl_overall": 0.02,                    # Overall inflation
    "Infl_e": 0.0231,                        # Inflation for fuel and electricity
    "LCOE": 0,                               # Levelized Cost of Energy, $/kWh
    "LCOE_el": 0,                            # Levelized Cost of Electricity, $/kWh
    "SCR": 0,                                # Self-consumption ratio
    "OPEX_el": 0,                            # Operating expenditures for LCOE_el
    "OPEX": 0,                               # Operating expenditures for LCOE
    "Den_LCOE": 0,                           # Denominator of LCOE
    "Den_LCOE_el": 0,                        # Denominator of LCOE_el
    "Q_ext_total": 0,                        # Total energy obtained from external boiler
    "P_grid_total": 0,                       # Total power obtained from grid
    "P_PV_total": 0,                         # Total power obtained from solar PV
    "Emission_factor_natural_gas": 180.7975, # Emission factor for natural gas, gCO2/kWh
    "Emission_factor_grid_elec": 304.14,     # Emission factor for grid electricity, gCO2eq/kWh
    "Emission_factor_PV_elec": 43,           # Emission factor for PV electricity, gCO2eq/kWh
    "Emissions_total_natural_gas": 0,        # Total emissions from natural gas, gCO2
    "Emissions_total_grid_elec": 0,          # Total emissions from grid electricity, gCO2eq
    "Emissions_total_PV_elec": 0,            # Total emissions from PV electricity, gCO2eq
    }

# %%
# This cell contains the functions that follow the energy management algorithm to calculate LCOE and LCOE_el

# Objective function to minimize LCOE
def objective_LCOE(vector):
    # The function argument vector is an array of Nominal PV capacity, HTES max. energy capacity, and max. generation capacity of PGU
        Model["WACC_real"] = (1 + Model["WACC_nom"])/(1 + Model["Infl_overall"]) - 1
        # Formula to compute Real WACC
        x = 0.99
        y = 0.1
        # x and y represent the charging condition for HTES and LTES required to evaluate certain parameters in the code

        for i in range(8760):
            Pnet[i] = Pc[i] - Pg[i]
            # The difference between power consumption and generation for a time interval
            Q_loss_LTES[i] = (a_LTES*np.sqrt(Model["E_LTES_max"]/0.018)*delta_T_LTES_E)/1000
            # LTES energy losses for a time interval
            Q_loss_HTES[i] = (a_HTES*np.sqrt(vector[1]/Ed_HTES)*delta_T_HTES)/1000 
            # HTES energy losses for a time interval
            if Pnet[i] < 0:
                # if Pnet is negative (power gen. > cons.), no power is obtained from grid (assuming PV energy can be directly supplied to the load or stored in HTES or LTES)
                P_grid[i] = 0
                P_out_PGU[i] = 0
                Q_loss_PGU[i] = 0
                Q_in_PGU[i] = 0
                Q_out_PGU[i] = 0
                Q_in_LTES[i] = 0
                if E_HTES[i] < vector[1]:
                    # HTES charging is prioritized over LTES: here, it is checked if it is fully charged or not
                    if Chh[i] >= 0:
                        # to check if Heat consumption>0 
                        if E_HTES[i] > x*vector[1]:
                            # to check if HTES charge is higher than 99%
                            P_in_HTES[i] = 0
                            # Power input to HTES = 0 as HTES is near to its capacity
                            if E_LTES[i] < y*Model["E_LTES_max"]:
                                # to check if LTES charge is lower than 10%
                                P_loss_PV[i] = 0
                                P_in_LTES[i] = -Pnet[i]
                                # LTES is charged if HTES is almost fully charged, and LTES if charges less than 10%
                            else:
                                P_loss_PV[i] = 0
                                # Lost PV electricity = 0
                                P_in_LTES[i] = 0
                                # Input power to LTES = 0 as it is almost fully charged
                        else:
                            P_in_HTES[i] = -Pnet[i]
                            # Power input to HTES is positive (as it is not fully charged) and equal to Pnet
                            P_loss_PV[i] = 0
                            # Power lost from PV = 0
                            P_in_LTES[i] = 0
                            # No power goes to LTES
                    else:
                        P_in_HTES[i] = -Pnet[i]
                        P_loss_PV[i] = 0
                        P_in_LTES[i] = 0
                else:
                    P_in_HTES[i] = 0
                    # HTES is at its max. capacity, thus it cannot store anymore
                    if E_LTES[i] > Model["E_LTES_max"]:
                        # to check is LTES is at its max capacity, thus it cannot store anymore, thus the extra energy produced by PV is lost
                        P_loss_PV[i] = -Pnet[i]
                        # Lost PV electricity = Pnet, as both HTES and LTES are fully charged
                        P_in_LTES[i] = 0
                        # Power input to LTES = 0
                    else:
                        P_loss_PV[i] = 0
                        # Power lost from PV = 0                                       
                        P_in_LTES[i] = -Pnet[i]
                        # Power input to LTES = Pnet, as it is not fully charged
            else:
                # Pnet is greater than or equal to zero (energy cons. >= gen.)
                P_loss_PV[i] = 0
                # Lost PV electricity = 0, thus, all of it goes to supply the electrical load
                P_in_LTES[i] = 0
                # LTES input power = 0
                P_in_HTES[i] = 0
                # HTES input power = 0
                
                if E_HTES[i] > Pnet[i]*Model["deltat"]/Model["Eff_PGU"]:
                    # To check if HTES has enough stored energy to supply
                    if Pnet[i] > vector[2]:
                        # To check if the electricity demand is higher than what PGU can supply
                        P_grid[i] = Pnet[i]-vector[2]
                        # If higher, take what you can from the grid
                        P_out_PGU[i] = vector[2]
                        # The rest of electricity is supplied from PGU
                    else:
                        P_grid[i] = 0
                        # No power is obtained from the grid as PGU can supply all that is required
                        P_out_PGU[i] = Pnet[i]
                        # PGU supplies Pnet
                else:
                    # HTES does not have enough stored energy
                    P_grid[i] = Pnet[i]
                    # Grid supplies Pnet
                    P_out_PGU[i] = 0
                    # No power is obtained from PGU
                Q_in_PGU[i] = P_out_PGU[i] / Model["Eff_PGU"]
                # Heat input to PGU ius calculated based on output and efficiency
                Q_out_PGU[i] = Q_in_PGU[i] - P_out_PGU[i]
                # Energy balance on PGU
                if E_LTES[i] > Model["E_LTES_max"]:
                    # To check if LTES is at its max. capacity, thus it cannot store anymore
                    Q_loss_PGU[i] = Q_out_PGU[i]
                    # Heat output of PGU is lost and not sent to LTES
                    Q_in_LTES[i] = 0
                    # Input power to LTES = 0
                else:
                    Q_loss_PGU[i] = 0
                    # Power lost from PGU = 0
                    Q_in_LTES[i] = Q_out_PGU[i]
                    # PGU output goes to LTES as it is not fully charged

            Ch[i] = Chh[i]

            if Ch[i] < ((E_LTES[i]/Model["deltat"])+Q_in_LTES[i]+P_in_LTES[i]-Q_loss_LTES[i]):
                # To check if the energy stored in LTES is enough to supply the heating load
                Qext[i] = 0
                # Heat obtained from the external boiler = 0, as LTES can satisfy the heating requirement
                Q_out_LTES[i] = Ch[i]
                # LTES output energy = the heating requirement  
            else:
                # LTES cannot supply the heating load
                Q_out_LTES[i] = E_LTES[i]/Model["deltat"]+Q_in_LTES[i]+P_in_LTES[i]-Q_loss_LTES[i]
                # Energy balance at LTES to calculate energy output=
                if Q_out_LTES[i] < 0:
                    Q_out_LTES[i] = 0
                    # To prevent LTES output from becoming negative
                Qext[i] = Ch[i]-Q_out_LTES[i]
                # Rest of the heating requirement is obtained from the external boiler

            if (E_LTES[i]+(P_in_LTES[i]+Q_in_LTES[i]-Q_out_LTES[i]-Q_loss_LTES[i])*Model["deltat"])<0:
                # If the above inequality is satisfied, it means that the LTES is unable to store energy for 
                # the ith interval, thus take the ith interval's energy from the external boiler, and keep
                # the ith interval unused LTES energy stored for the next, (i+1)th interval.
                E_LTES[i+1] = E_LTES[i]
                P_in_LTES[i] = 0
                Q_in_LTES[i] = 0
                Q_out_LTES[i] = 0
                Q_loss_LTES[i] = 0
                Qext[i] = Ch[i]

            else:
                E_LTES[i+1] = E_LTES[i]+(P_in_LTES[i]+Q_in_LTES[i]-Q_out_LTES[i]-Q_loss_LTES[i])*Model["deltat"]
                # Energy balance at LTES to calculate the stored energy for the (i+1)th interval

            if (E_HTES[i]+(P_in_HTES[i]-Q_in_PGU[i]-Q_loss_HTES[i])*Model["deltat"])<0:
                # If the above inequality is satisfied, it means that the HTES is unable to store energy for 
                # the ith interval, thus take the ith interval's energy from the grid, and keep
                # the ith interval unused HTES energy stored for the next, (i+1)th interval.
                E_HTES[i+1] = E_HTES[i]
                P_in_HTES[i] = 0
                Q_in_PGU[i] = 0
                Q_loss_HTES[i] = 0
                Q_out_PGU[i] = 0
                P_out_PGU[i] = 0
                Q_loss_PGU[i] = 0

                P_grid[i] = Pc[i] - ( (Pg[i] - (P_in_LTES[i]+P_loss_PV[i]+P_in_HTES[i])) + P_out_PGU[i])
                # Power obtained from grid is calculated from the power balance equation based on updated values
                
                if P_grid[i] < 0:
                    # To check if Pgrid is negative
                    P_loss_PV[i] = ( (Pg[i] - (P_in_LTES[i]+Pc[i]+P_in_HTES[i])) + P_out_PGU[i])
                    # Rest of the electricity is obtained from solar PV
                    P_grid[i] = 0
                    # Pgrid is turned zero if it were negative

                Q_in_LTES[i] = 0
                # Heat input to LTEs = 0

            else:
                E_HTES[i+1] = E_HTES[i]+(P_in_HTES[i]-Q_in_PGU[i]-Q_loss_HTES[i])*Model["deltat"]
                # Energy balance at HTES to calculate the stored energy for the (i+1)th interval

            Q_loss[i] = Q_loss_HTES[i]+Q_loss_LTES[i]+Q_loss_PGU[i]+P_loss_PV[i]
            # Total energy losses are computed as a sum of losses from HTEs, LTES, PGU, and PV

        Model["CAPEX"] = vector[0]*Model["CAPEX*_PV"] + vector[1]*Model["CAPEX*_HTES"] + Model["E_LTES_max"]*Model["CAPEX*_LTES"]+ vector[2]*Model["CAPEX*_PGU"] + (Cc_max/COP_EHP)*Model["CAPEX*_EHP"]
        # Total capital expenditure, in $ is computed as the sum of the product of capital cost factors and energy capacities of PV, HTES, LTE, PGU, and EHP
        Term_1 = np.sum(P_grid)
        # Total electricity obtained from grid throughout the year
        Term_2 = np.sum(Pc)
        # Total electricity consumed by load throughout the year
        Term_3 = np.sum(Qext)
        # Total heat obtained from external boiler (running on natural gas) throughout the year
        Term_4 = np.sum(Pc+Ch)
        # Total energy requirements (heat + electricity) throughout the year

        def func_OPEX(x):
            # This function calculates the operating expenditure (normalized by WACC_nom) for LCOE_el
            # It consists of fixed and variable cost of grid electricity (both multiplied by electricity inflation rate)
            return ((pow((1+Model["Infl_e"]),x)*(Model["OPEX_elec_fix"]*max(P_grid)+Model["OPEX_elec_var"]*Term_1)/(pow((1+Model["WACC_nom"]),x))))
        
        def func_OPEX_2(x):
            # This function calculates the operating expenditure (normalized by WACC_nom) for LCOE
            # It consists of fixed and variable cost of grid electricity and natural gas (both multiplied by electricity and fuel inflation rate)
            return ((pow((1+Model["Infl_e"]),x)*(Model["OPEX_elec_fix"]*max(P_grid)+Model["OPEX_elec_var"]*Term_1)/(pow((1+Model["WACC_nom"]),x))+\
            pow((1+Model["Infl_e"]),x)*(Model["OPEX_fuel_fix"]+Model["OPEX_fuel_var"]*Term_3)/(pow((1+Model["WACC_nom"]),x))))
        
        def func_Den(x):
            # This function calculates the denominator of LCOE: yearly electric demand normalized by WACC_real
            return (Term_2/(pow((1+Model["WACC_real"]),x)))
        
        def func_Den_2(x):
            # This function calculates the denominator of LCOE: yearly energy (heat + electricity) demand normalized by WACC_real
            return (Term_4/(pow((1+Model["WACC_real"]),x)))
        
        L = np.arange(1,Model["Lifetime"]+1,1)

        Model["OPEX_el"]=np.sum(func_OPEX(L))
        # This function calculates the operating expenditures (for LCOE_el) for the entire lifetime of the project
        Model["LCOE_el"] = (Model["CAPEX"] + Model["OPEX_el"])/(np.sum(func_Den(L)))
        # LCOE_el ($/kWh) is computed according to its definition
        Model["OPEX"] = np.sum(func_OPEX_2(L))
        # This function calculates the operating expenditures (for LCOE) for the entire lifetime of the project
        Model["Den_LCOE"] = np.sum(func_Den_2(L))
        Model["LCOE"] = (Model["CAPEX"] + Model["OPEX"])/(Model["Den_LCOE"])
        # LCOE ($/kWh) is computed according to its definition

        Model["Q_ext_total"] = Term_3
        Model["P_grid_total"] = Term_1
        Term_5 = np.sum(Pg - (P_in_LTES+P_loss_PV+P_in_HTES))
        Model["P_PV_total"] = Term_5
        
        Model["Emissions_total_natural_gas"] = Model["Emission_factor_natural_gas"] * Model["Q_ext_total"]/1e6
        # Calculation of total emissions from natural gas in Metric tons CO2/year
        Model["Emissions_total_grid_elec"] = Model["Emission_factor_grid_elec"] * Model["P_grid_total"]/1e6
        # Calculation of total emissions from ngrid electricity in Metric tons CO2/year
        Model["Emissions_total_PV_elec"] = Model["Emission_factor_PV_elec"] * Model["P_PV_total"]/1e6
        # Calculation of total emissions from PV electricity in Metric tons CO2/year

        Model["Emissions_total_no_energy_storage"] = (Model["Emission_factor_natural_gas"]*np.sum(Ch)+Model["Emission_factor_grid_elec"]*sum(Pc))/1e6
        # Calculation of total emissions (for no energy storage) in Metric tons CO2/year

        return Model["LCOE"]

# Objective function to minimize LCOE_el
def objective_LCOE_el(vector):
        Model["WACC_real"] = (1 + Model["WACC_nom"])/(1 + Model["Infl_overall"]) - 1
        x = 0.99
        y = 0.1
        for i in range(8760):
            Pnet[i] = Pc[i] - Pg[i]
            Q_loss_LTES[i] = (a_LTES*np.sqrt(Model["E_LTES_max"]/0.018)*delta_T_LTES_E)/1000
            Q_loss_HTES[i] = (a_HTES*np.sqrt(vector[1]/Ed_HTES)*delta_T_HTES)/1000
            if Pnet[i]<0:
                P_grid[i] = 0
                P_out_PGU[i] = 0
                Q_loss_PGU[i] = 0
                Q_in_PGU[i] = 0
                Q_out_PGU[i] = 0
                Q_in_LTES[i] = 0
                if E_HTES[i]<vector[1]:
                    if Chh[i]>=0:                                        
                        if E_HTES[i]>x*vector[1]:                        
                            P_in_HTES[i]=0                               
                            if E_LTES[i]<y*Model["E_LTES_max"]:          
                                P_loss_PV[i]=0
                                P_in_LTES[i]=-Pnet[i]
                            else:
                                P_loss_PV[i]=0                           
                                P_in_LTES[i]=0
                        else:
                            P_in_HTES[i]=-Pnet[i]
                            P_loss_PV[i]=0
                            P_in_LTES[i]=0
                    else:
                        P_in_HTES[i]=-Pnet[i]
                        P_loss_PV[i]=0
                        P_in_LTES[i]=0
                else:                                                    
                    P_in_HTES[i]=0
                    if E_LTES[i]>Model["E_LTES_max"]:                    
                        P_loss_PV[i]=-Pnet[i]                            
                        P_in_LTES[i]=0
                    else:
                        P_loss_PV[i]=0                                   
                        P_in_LTES[i]=-Pnet[i]
            else:                                                        
                P_loss_PV[i] = 0
                P_in_LTES[i] = 0
                P_in_HTES[i] = 0
                if E_HTES[i]>Pnet[i]*Model["deltat"]/Model["Eff_PGU"]:   
                    if Pnet[i]>vector[2]:                                
                        P_grid[i]=Pnet[i]-vector[2]
                        P_out_PGU[i]=vector[2]
                    else:
                        P_grid[i]=0
                        P_out_PGU[i]=Pnet[i]
                else:
                    P_grid[i]=Pnet[i]
                    P_out_PGU[i]=0
                Q_in_PGU[i]=P_out_PGU[i]/Model["Eff_PGU"]
                Q_out_PGU[i]=Q_in_PGU[i]-P_out_PGU[i]
                if E_LTES[i]>Model["E_LTES_max"]:
                    Q_loss_PGU[i]=Q_out_PGU[i]
                    Q_in_LTES[i]=0
                else:
                    Q_loss_PGU[i]=0
                    Q_in_LTES[i]=Q_out_PGU[i]
            Ch[i] = Chh[i]                       
            if Ch[i]<((E_LTES[i]/Model["deltat"])+Q_in_LTES[i]+P_in_LTES[i]-Q_loss_LTES[i]):
                Qext[i]=0
                Q_out_LTES[i]=Ch[i]        
            else:
                Q_out_LTES[i] = E_LTES[i]/Model["deltat"]+Q_in_LTES[i]+P_in_LTES[i]-Q_loss_LTES[i]
                if Q_out_LTES[i]<0:
                    Q_out_LTES[i]=0
                Qext[i] = Ch[i]-Q_out_LTES[i]
            if (E_LTES[i]+(P_in_LTES[i]+Q_in_LTES[i]-Q_out_LTES[i]-Q_loss_LTES[i])*Model["deltat"])<0:
                E_LTES[i+1]=E_LTES[i]
                P_in_LTES[i] = 0
                Q_in_LTES[i] = 0
                Q_out_LTES[i] = 0
                Q_loss_LTES[i] = 0
                Qext[i] = Ch[i]
            else:
                E_LTES[i+1] = E_LTES[i]+(P_in_LTES[i]+Q_in_LTES[i]-Q_out_LTES[i]-Q_loss_LTES[i])*Model["deltat"]
            if (E_HTES[i]+(P_in_HTES[i]-Q_in_PGU[i]-Q_loss_HTES[i])*Model["deltat"])<0:
                E_HTES[i+1]=E_HTES[i]
                P_in_HTES[i] = 0
                Q_in_PGU[i] = 0
                Q_loss_HTES[i] = 0
                Q_out_PGU[i] = 0
                P_out_PGU[i] = 0
                Q_loss_PGU[i] = 0
                P_grid[i] = Pc[i] - ( (Pg[i] - (P_in_LTES[i]+P_loss_PV[i]+P_in_HTES[i])) + P_out_PGU[i])
                if P_grid[i]<0:
                    P_loss_PV[i] = ( (Pg[i] - (P_in_LTES[i]+Pc[i]+P_in_HTES[i])) + P_out_PGU[i])
                    P_grid[i] = 0
                Q_in_LTES[i] = 0
            else:
                E_HTES[i+1] = E_HTES[i]+(P_in_HTES[i]-Q_in_PGU[i]-Q_loss_HTES[i])*Model["deltat"]    
            Q_loss[i] = Q_loss_HTES[i]+Q_loss_LTES[i]+Q_loss_PGU[i]+P_loss_PV[i]
        Model["CAPEX"] = vector[0]*Model["CAPEX*_PV"] + vector[1]*Model["CAPEX*_HTES"] + Model["E_LTES_max"]*Model["CAPEX*_LTES"]+ vector[2]*Model["CAPEX*_PGU"] + (Cc_max/COP_EHP)*Model["CAPEX*_EHP"]
        Term_1 = np.sum(P_grid)
        Term_2 = np.sum(Pc)
        Term_3 = np.sum(Qext)
        Term_4 = np.sum(Pc+Ch)
        def func_OPEX(x):
            return ((pow((1+Model["Infl_e"]),x)*(Model["OPEX_elec_fix"]*max(P_grid)+Model["OPEX_elec_var"]*Term_1)/(pow((1+Model["WACC_nom"]),x))))
        def func_OPEX_2(x):
            return ((pow((1+Model["Infl_e"]),x)*(Model["OPEX_elec_fix"]*max(P_grid)+Model["OPEX_elec_var"]*Term_1)/(pow((1+Model["WACC_nom"]),x))+\
            pow((1+Model["Infl_e"]),x)*(Model["OPEX_fuel_fix"]+Model["OPEX_fuel_var"]*Term_3)/(pow((1+Model["WACC_nom"]),x))))
        def func_Den(x):
            return (Term_2/(pow((1+Model["WACC_real"]),x)))
        def func_Den_2(x):
            return (Term_4/(pow((1+Model["WACC_real"]),x)))
        L = np.arange(1,Model["Lifetime"]+1,1)
        Model["OPEX_el"]=np.sum(func_OPEX(L))
        Model["Den_LCOE_el"] = np.sum(func_Den(L))
        Model["LCOE_el"] = (Model["CAPEX"] + Model["OPEX_el"])/(Model["Den_LCOE_el"])
        Model["Q_ext_total"] = Term_3
        Model["P_grid_total"] = Term_1
        Term_5 = np.sum(Pg - (P_in_LTES+P_loss_PV+P_in_HTES))
        Model["P_PV_total"] = Term_5
        Model["Emissions_total_natural_gas"] = Model["Emission_factor_natural_gas"] * Model["Q_ext_total"]/1e6
        Model["Emissions_total_grid_elec"] = Model["Emission_factor_grid_elec"] * Model["P_grid_total"]/1e6
        Model["Emissions_total_PV_elec"] = Model["Emission_factor_PV_elec"] * Model["P_PV_total"]/1e6
        return Model["LCOE_el"]

# %%
# This cell calculates and optimizes LCOE and LCOE_el, and plots them and CO2eq. emissions

# Initial guess for optimization
initial_guess  = [9.5, 40, 1]

# Constraints for optimization
cons = ({'type': 'ineq', 'fun': lambda x: Model["P_nom_PV"]},
        {'type': 'ineq', 'fun': lambda x: Model["E_HTES_max"]},
        {'type': 'ineq', 'fun': lambda x: Model["Pmax_PGU"]},
        )

# Bounds for optimization
bnds = ((5, 12), (15, 80), (0.5,3))

LCOE_el_base = objective_LCOE_el([Model["P_nom_PV"], Model["E_HTES_max"], Model["Pmax_PGU"]])
# To store the base-case value of LCOE_el before optimization

#The code below is to plot the total emissions figure
A = Model["CAPEX"]/Model["Den_LCOE_el"]
B = Model["OPEX_el"]/Model["Den_LCOE_el"]
NG_1 = Model["Emissions_total_natural_gas"]
GE_1 = Model["Emissions_total_grid_elec"]
PV_1 = Model["Emissions_total_PV_elec"]
result = minimize(objective_LCOE_el, initial_guess, method='Nelder-Mead', bounds=bnds, constraints= cons, tol=10e-5)

NG_2 = Model["Emissions_total_natural_gas"]
GE_2 = Model["Emissions_total_grid_elec"]
PV_2 = Model["Emissions_total_PV_elec"]

C = Model["CAPEX"]/Model["Den_LCOE_el"]
D = Model["OPEX_el"]/Model["Den_LCOE_el"]
NG_3 = Model["Emission_factor_natural_gas"]*np.sum(Ch)/1e6
GE_3 = Model["Emission_factor_grid_elec"]*sum(Pc)/1e6
PV_3 = 0

solution = result['x']
LCOE_el_optimized = objective_LCOE_el(solution)

LCOE_base = objective_LCOE([Model["P_nom_PV"], Model["E_HTES_max"], Model["Pmax_PGU"]])
E = Model["CAPEX"]/Model["Den_LCOE"]
F = Model["OPEX"]/Model["Den_LCOE"]
result = minimize(objective_LCOE, initial_guess, method='Nelder-Mead', bounds=bnds, constraints= cons, tol=10e-5)
G = Model["CAPEX"]/Model["Den_LCOE"]
H = Model["OPEX"]/Model["Den_LCOE"]
solution = result['x']
LCOE_optimized = objective_LCOE(solution)

species = (
    "a. LCOE$_{el}$ (Base-case)",
    "b. LCOE$_{el}$ (Optimized)",
    "c. LCOE (Base-case)",
    "d. LCOE (Optimized)"
)

weight_counts = {
    "Capital cost": np.array([A, C, E, G]),
    "Operating cost": np.array([B, D, F, H]),
}
width = 0.5

fig, ax = plt.subplots()
bottom = np.zeros(4,dtype='float64')

for boolean, weight_count in weight_counts.items():
    p = ax.bar(species, weight_count, width, label=boolean, bottom=bottom)
    np.add(bottom, weight_count, out=bottom, casting="unsafe")

for container in ax.containers:
    ax.bar_label(container,fmt='%.3f',fontsize=20)
fig.set_size_inches(12,7)
plt.ylabel("\$/kWh", size=24)
plt.xticks(fontsize=24, rotation=10)
plt.yticks(fontsize=22)
plt.tight_layout(pad=0)
ax.legend(bbox_to_anchor=(1.1, 1.05))
plt.legend(fontsize=24)
plt.savefig('LCOE_and_LCOE_el_optimization.jpeg', dpi=300, bbox_inches='tight', pad_inches=0)
plt.tight_layout()
plt.show()

species = (
    "a. No energy storage",
    "b. Base-case",
    "c. Optimized",
)

weight_counts = {
    "Natural gas": np.array([NG_3, NG_1, NG_2]),
    "Grid electricity": np.array([GE_3, GE_1, GE_2]),
    "PV electricity": np.array([PV_3, PV_1, PV_2]),
}
width = 0.5

fig, ax = plt.subplots()
bottom = np.zeros(3,dtype='float64')

for boolean, weight_count in weight_counts.items():
    p = ax.bar(species, weight_count, width, label=boolean, bottom=bottom)
    np.add(bottom, weight_count, out=bottom, casting="unsafe")

for container in ax.containers:
    ax.bar_label(container,fmt='%.3f',fontsize=20,label_type="center")
fig.set_size_inches(10,7)
plt.ylabel("CO$_{2}$eq. emissions (tons/year)", size=24)
plt.xticks(fontsize=24)
plt.yticks(fontsize=22)
plt.ylim(top=10.5)
plt.tight_layout(pad=0)
ax.legend(bbox_to_anchor=(1.1, 1.05))
plt.legend(fontsize=24)
plt.savefig('CO2_emissions.jpeg', dpi=300, bbox_inches='tight')
plt.tight_layout()
plt.show()